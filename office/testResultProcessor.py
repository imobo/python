import matplotlib.pyplot as plt
from matplotlib import gridspec
import os
import copy
import datetime


def readExcel(file):
    '''
    input:
    excel file name of the source data
    output:
    dataSum = {challenger name:{reviewer name:{pts:,opinions:,lables:}} * reviewer} * challenger
    '''
    from openpyxl import load_workbook
    
    sCol = 3
    sRow = 2
    dataSum = {}
    
    wb = load_workbook(file)
    names = wb.sheetnames
    for n in names:
        ws = wb[n]
        countR = ws.max_row + 1
        countC = ws.max_column + 1

        data = {}
        for c in range(sCol, countC):
            temp1, temp2, tempLab = [], [], []
            for r in range(sRow, countR):
                val = ws.cell(row = r, column = c).value
                # turn "-""nashi" into None
                val = dataNormalise(val)
                # get label of every rows
                subKey = ws.cell(row = r, column = sCol - 1).value
                if len(tempLab) < countR - sRow:
                    tempLab.append(subKey)
                    
                # temp1 for pts, temp2 for comment
                if isNum(val) == True:
                    temp1.append(val)
                else:
                    if val != None:
                        temp2.append(val)
                        
            # if not full mark, turn temp1 to None
            if len(temp1)<=1:
                temp1 = None

            if temp1 == None:
                tempLab = None
            else:
                for i in range(len(tempLab), len(temp1), -1):
                    tempLab.pop(i - 1)
            
            temp = {'pts':temp1, "opinion":temp2, "labels":tempLab}
            key = ws.cell(row = sRow - 1, column = c).value
            data[key] = temp
            
        dataSum[n] = data
        
    return dataSum


def dataNormalise(strings):
    marks = ['-',"无","なし"]
    if isNum(strings) == False:
        for m in marks:
            if strings == m:
                return None
    return strings
        

def isNum(var):
    if isinstance(var, int) == True:
        return True
    elif isinstance(var, float) == True:
        return True
    else:
        return False


def isXlFile(file):
    extension = ["xlsx", "xlsm"]
    if file[-4:] in extension:
        return True
    else:
        return False


def addSeparator(n):
    if n[-2:] != "//":
        d = n + "//"
    else:
        d = n
    return d


def saveDictToTxt(dir, dict, newName):        
    f = open(dir + newName + ".txt","w",encoding='utf-8')
    f.write( str(dict) )
    f.close()


def readTxt(dir,name):
    f = open(dir + name + ".txt","r",encoding='utf-8')
    content = f.readlines()
    content = eval(content[0])
    f.close()
    return content


def extractData(d, k):
    p, c, l = [], [], []
    # k: keyword(challenger)  d: dictionary contain every data of every challenger
    # temp : data of a specific challenger
    temp = d[k]
    # loop every reviewer in list
    for i in temp.keys():
        reviewer = i
        # get points and commets from multiple reviewer
        for j in temp[reviewer].keys():
            if j == "pts":
                p.append(temp[reviewer][j])
            elif j == "opinion":
                c.append(temp[reviewer][j])
            # get labels in dict, only extract one time
            elif j == "labels" and len(l) < 1:
                l.append(temp[reviewer][j])
    return p, c, l[0]


def getListAvg(lt):
    """
    [[reviewer 1],[reviewer 2],[reviewer 3]]  -->> [[avg of all 3 reviewer],[max value of ],[minimum value of ]]
    """
    l = copy.deepcopy(lt)
    # detete element dont have full mark
    for i in range(len(l)-1,0,-1):
        if l[i] == None:
            l.pop(i)
        elif len(l[i]) < 2:
            l.pop(i)
            
    # create a empty list have value 0 as element
    temp = [0] * len(l[0])
    lmax = [0] * len(l[0])
    lmin = [99] * len(l[0])
    
    for i in range(0,len(l)):
        # sum up
        for j in range(0,len(temp)):
            temp[j] = temp[j] + l[i][j]
            if l[i][j] > lmax[j]:
                lmax[j] = l[i][j]
            if l[i][j] < lmin[j]:
                lmin[j] = l[i][j]
    # get average
    for i in range(0,len(temp)):
        temp[i] = temp[i] / len(l)

    return temp, lmax, lmin


def dimensionReduction(arr):
    arr = [x for y in arr for x in y]
    return arr


def factorListCompare(l, f):
    temp = None
    r = []
    for i in range(0,len(f)):
        temp = dimensionReduction(f[i])
        if len(l) == len(temp):
            r = f[i] #r = temp
            return r
    if r == []:
        print("...Did not find a factor list have the same length")
        exit


def listPreprocessForPlot(lavg, lmax, lmin, fList):
    """
    [[1,2,3],[1,2,3]]  -->> [[1,1],[2,2][3,3]]
    """
    ff = factorListCompare(lavg, fList)
    ff = dimensionReduction(ff)
    
    l = []
    for i in range(0, len(lavg)):
        temp = []
        temp.append(lavg[i] * ff[i])
        temp.append(lmax[i] * ff[i])
        temp.append(lmin[i] * ff[i])
        l.append(temp)
        
    return l


def plotBox(ls, n, l, p, y, fList, norBase1, norBase2, com):                                       # TODO: ADD COMMENT?
    """
    ls = [[avg,max,min],[avg,max,min], ...
    n = challenger name (to generate file name)
    l = labels
    p = save path
    y = year (to generate file name)
    fList = the list of factor
    norBase = max and min data for normalise
    """
    
    xf = 1.2
    yf = 6
    s = 0.5
    f = factorListCompare(ls, fList)

    c = 0
    tempCom = "Comments: \n"
    for i in range(0, len(com)):
        if com[i] != []:
            for j in range(0, len(com[i])):
                tempCom = tempCom + com[i][j] + "\n"
    commentSpace = tempCom.count("\n") * 0.2
    
    partition = [len(f[0]) + s, len(ls) - s]
    n = n + ' - 资格审查结果' + "(" + str(y) +")"
    
    # 设置简黑字体
    plt.rcParams['font.sans-serif'] = ['SimHei']
    # 解决‘-’bug
    plt.rcParams['axes.unicode_minus'] = False
    # 设置画布的尺寸
    plt.figure(figsize=(len(ls) * xf, yf + commentSpace))
    # 标题，并设定字号大小
    plt.title(n, fontsize = len(ls) * xf)
    # 图例
    labels = l
    # 图表1
    if commentSpace > 5:
        gSize = 3
    else:
        gSize = 2
    gs = gridspec.GridSpec(gSize,1)
    ax1 = plt.subplot(gs[:1,0])
    # 分割线
    ax1.vlines(partition[0], 0, yf, colors = "b", linestyles = "dashed")
    ax1.vlines(partition[1], 0, yf, colors = "b", linestyles = "dashed")
    # grid=False：代表不显示背景中的网格线
    ax1.boxplot(ls, labels = labels, showmeans = True)
    
    # min-max normalization
    # fromula: nex_x = ( x - min(X) ) / ( max(X) - min(X) )
    # get the avg of ls
    temp = []
    for i in range(0, len(ls)):
        temp.append(ls[i][0])

    tempf = None
    flag= 0
    for i in range(0, len(fList)):
        if len(dimensionReduction(fList[i])) == len(temp):
            flag = i
            tempf = copy.deepcopy(fList[i])

    elms = elmCompress(temp, tempf)
    
    for i in range(0,len(elms)):
        # [3,2,3,1,1]
        tempf[i] = divideElemByOne(tempf[i])
        tempSum2 = 0
        for j in range(0,len(tempf[i])):
            tempSum1 = 0
            tempSum1 = elms[i] * tempf[i][j]
            tempSum2 = tempSum2 + tempSum1
        maxx = norBase1[flag][i]
        minx = norBase2[flag][i]
        elms[i] = tempSum2 / len(tempf[i])
        elms[i] = (elms[i] - minx) / (maxx - minx)
        
    # fontsize of text
    fontsize = len(ls) * xf / 1.5
    # add text
    temps = s
    for i in range(0, len(tempf) - 1):
        if i == 0:
            plt.text(temps, yf - s, " 顺位值:" + str(round(elms[i],2)), fontsize = fontsize)
        temps = temps + len(tempf[i])
        plt.text(temps, yf - s, " 顺位值:" + str(round(elms[i + 1],2)), fontsize = fontsize)
    # y轴上下限
    plt.ylim([0, yf])
    ax2 = plt.subplot(gs[gSize-1,0])
    plt.axis('off')
    # 表格行名
    ax2.annotate(tempCom, xy=(0, 0))
    
    # 显示/保存图像
    plt.savefig(n + ".png" )
    # plt.show()
    print(n, "has been saved in", p)


def getNormaliseBase(d, fList):                     #分管理和技术，f中的每一组分开取得
    '''
    OUTPUT: rMax[[part1 pts, part2 pts, part3 pts],[........]]
                    (first set of factors)          (second set of factors)

    '''
    rMax,rMin = [],[]
    # for every set of factor
    for f in range(0,len(fList)):
        # get factor in a 1d array
        dr = dimensionReduction(fList[f])
        # bMax[number of factor]: the max, bMin[...]: the mininum
        bMax, bMin = [0] * len(fList[f]), [99] * len(fList[f])
        # get challenger data in dict one by one
        for k in d.keys():
            # p: set of points of challenger k from reviewers [1,2,3,4,5,6...] * 5
            p, c, l = extractData(d, k)
            tempList = [0] * len(dr)
            emptyCount = 0
            # for every set of points               i = reviewer    j = point
            for i in range(0, len(p)):
                if p[i] != None and len(p[i]) == len(dr):
                    for j in range(0, len(p[i])):
                        tempList[j] = tempList[j] + p[i][j]
                if p[i] == None:
                    emptyCount = emptyCount + 1

            # igrone list have default 0 value
            # occured when ?
            if sum(tempList) != 0:
                # get avg of reviewers
                for i in range(0, len(tempList)):
                    tempList[i] = tempList[i] / (len(p) - emptyCount)
                    
                # compress elements(pts) to the number of factor set
                elms = elmCompress(tempList, fList[f])
                for j in range(0, len(elms)):
                    if elms[j] > bMax[j]:
                        bMax[j] = elms[j]
                    if elms[j] < bMin[j]:
                        bMin[j] = elms[j]
                        
        rMax.append(bMax)
        rMin.append(bMin)
    #print(rMax, rMin)
    return rMax, rMin


def elmCompress(tempList, f):
    count = 0
    l = []
    for i in range(0, len(f)):
        temp = 0
        for j in range(0, len(f[i])):
            temp = temp + tempList[count]
            count = count + 1
        temp = temp / len(f[i])
        l.append(temp)
    return l


def divideElemByOne(l):
    """
    [1,2,3,4,5] = [1.0, 0.5, 0.33, 0.25, 0.2]
    """
    newL = [0] * len(l)
    for i in range(0,len(l)):
        newL[i] = 1 / l[i]
    return newL


def main():
    newName = "result"
    path = os.path.dirname(__file__)
    path = addSeparator(path)

    # normal
    f1 = [[1,1,1,1,1,1],[0.3333,0.5,0.3333,1,1],[0.25]]
    # tech
    f2 = [[1,1,1,1,1,1],[0.25,0.25,0.5],[0.25]]
    
    ffff = [f1, f2]
    
    inp = int(input("Select a function: ").strip())
    # output results into a text file

    
    if inp == 1:
        print("This function will import data from all the excel file in the same directory.")
        dataYear = str(input("What is the year of the data: ").strip())
        # get every file in the smae directory
        files = os.listdir(path)
        ttlData = {}
        # loop every file
        for file in files:
            # check if a file is a excel file
            if isXlFile(file) == True:
                print("...processing to text file ",file)
                # get data in the Excel file and update it into dictionary(ttlData)
                ttlData.update(readExcel(path + file))
        # save as txt file
        saveDictToTxt(path, ttlData, dataYear + newName)
        print("...result save as : ", dataYear + newName, ".txt")


    # output results of specificed challengers in multiple years
    elif inp == 2:
        startYear = int(input("Enter an year to start import process: ").strip())
        endYear = int(datetime.datetime.today().year) + 1
        challengers = input("Input the name of challengers separate by , (tpye all to get data of everyone): ")
        challengers = challengers.split(",")
        
        records = []
        # get data in all years 
        for year in range(startYear,endYear):
            try:
                # store data in record. record[0] = data in first year, record[1] data in second year etc.
                records.append(readTxt(path, str(year) + newName))
                print("...data import successfully in", year)
                # get input is "all", get record of every challenger in dictionary
                if challengers == ['all']:
                    # reset variable to a empty list
                    challengers = []
                    # get data from the last element of records
                    for challenger in records[len(records)- 1].keys():
                        challengers.append(challenger)
                # loop and output data for specificed people                    
                for challenger in challengers:
                    try:
                        pts, comment, labels = extractData(records[len(records)- 1], challenger)
                        avg, mx, mn = getListAvg(pts)
                        rearrange = listPreprocessForPlot(avg, mx, mn, ffff)
                        gnb1,gnb2 = getNormaliseBase(records[len(records)- 1], ffff)
                        plotBox(rearrange, challenger, labels, path, year, ffff, gnb1, gnb2, comment)

                    except KeyError:
                        print("Did not find any record of",challenger, "in", year)
                        
            except FileNotFoundError:
                print("...data import failure in", year)
                pass


    else:
        pass
    

if __name__ == '__main__':
    main()
